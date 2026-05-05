#!/usr/bin/env python3
"""
import_legacy.py — One-time import of the legacy CanvasCircle data into Supabase.

Reads:
  - ../Artwork_Submission_Form.xlsx   (the exported Google Sheet)
  - ../catalog_images/                (one image file per listing)

Writes (via the Supabase service-role key — bypasses RLS):
  1. Wipes the four test seed listings (the 1111…/2222…/3333…/4444… UUIDs).
  2. Creates an auth user for any seller email that doesn't have one.
     (The handle_new_user trigger then auto-creates their profile.)
  3. Updates each profile with display_name, location, post_header,
     post_footer, facebook_profile_url derived from the spreadsheet.
  4. For each row in the spreadsheet:
       - Generates a fresh UUID for the listing.
       - Uploads the image to listing-images/{uuid}/0-{filename}.
       - Inserts the listing row + its listing_images row.

Idempotent: re-running won't create duplicates because we delete-then-insert
the listings in one pass. (Sellers already created are skipped; profiles are
upserted; bucket uploads use upsert=true so they overwrite cleanly.)

Setup:
    cd CanvasCircle_2
    python3 -m venv .venv && source .venv/bin/activate
    pip install -r scripts/requirements.txt

    # Get this from Supabase > Project Settings > API > service_role key.
    # NEVER commit it. NEVER paste it in chat. Treat it like a root password.
    export SUPABASE_SERVICE_ROLE_KEY="eyJ..."
    export SUPABASE_URL="https://xwieomjsqwcswoadrvkv.supabase.co"

    python scripts/import_legacy.py

Run order matters: db/migrations/001_artwork_category_and_framed_size.sql
must already have been run in the Supabase SQL editor before this script.
"""

from __future__ import annotations

import os
import sys
import uuid
import mimetypes
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from supabase import create_client, Client

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://xwieomjsqwcswoadrvkv.supabase.co")
SERVICE_KEY  = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")

if not SERVICE_KEY:
    sys.exit(
        "ERROR: SUPABASE_SERVICE_ROLE_KEY env var is not set.\n"
        "Get it from Supabase > Project Settings > API > service_role key,\n"
        "then run:  export SUPABASE_SERVICE_ROLE_KEY='eyJ...'"
    )

REPO_ROOT  = Path(__file__).resolve().parent.parent
XLSX_PATH  = REPO_ROOT / "Artwork_Submission_Form.xlsx"
IMG_DIR    = REPO_ROOT / "catalog_images"
BUCKET     = "listing-images"

# The four test seed UUIDs from db/seed.sql — wipe these before importing.
TEST_SEED_UUIDS = [
    "11111111-1111-1111-1111-111111111111",
    "22222222-2222-2222-2222-222222222222",
    "33333333-3333-3333-3333-333333333333",
    "44444444-4444-4444-4444-444444444444",
]

# Map legacy XLSX column index → meaning. Source of truth:
# the headers we printed during exploration.
COL = {
    "timestamp": 0, "seller_name": 1, "artist_name": 2, "artwork_title": 3,
    "medium": 4, "category": 5, "size": 6, "framed_size": 7, "price": 8,
    "shipping": 9, "coa": 10, "description": 11, "image_url_form": 12,
    "seller_email_pref": 13, "seller_email_alt": 14,
    "listing_id": 16, "image_file": 17, "source_image_url": 18,
    "status": 19, "seller_token": 20, "seller_profile_url": 21,
    "updated_at": 22, "previous_price": 23, "price_updated_at": 24,
    "moderation_status": 25, "management_link_sent_at": 26,
    "management_link_last_sent_to": 27,
    "seller_post_header": 28, "seller_post_footer": 29,
    "seller_location": 30, "seller_mood": 31,
    "last_renewed_at": 34, "renewal_warning_sent_at": 35,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _to_float(token: str) -> Optional[float]:
    """Parse '12', '12.5', '12 1/8', '1/2'. Returns None on failure."""
    t = (token or "").strip()
    if not t:
        return None
    # Strip stray quotes/inch marks and trailing punctuation
    for ch in ('"', "”", "“", "'", "’", "‘"):
        t = t.replace(ch, "")
    t = t.strip().rstrip(".,")
    if not t:
        return None
    # Mixed fraction "12 1/8"
    parts = t.split()
    try:
        if len(parts) == 2 and "/" in parts[1]:
            whole = float(parts[0])
            num, den = parts[1].split("/")
            return whole + float(num) / float(den)
        if len(parts) == 1 and "/" in parts[0]:
            num, den = parts[0].split("/")
            return float(num) / float(den)
        return float(t)
    except (ValueError, ZeroDivisionError):
        return None


def parse_dimensions(s: Optional[str]):
    """'20 x 20' → (20, 20, None). '14 x 6 x 6' → (14, 6, 6).
    Handles inch marks (", ”), mixed fractions (12 1/8)."""
    if not s:
        return (None, None, None)
    cleaned = str(s).replace("×", "x").replace("X", "x")
    parts = [p for p in cleaned.split("x")]
    nums = [_to_float(p) for p in parts]
    while len(nums) < 3:
        nums.append(None)
    return tuple(nums[:3])


def normalize_email(s: Optional[str]) -> Optional[str]:
    if not s:
        return None
    return str(s).strip().lower()


def iso_or_none(v):
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        return v.isoformat()
    return str(v)


def find_image_file(image_file_name: str) -> Optional[Path]:
    """Some rows reference .jpeg, the actual file may be .jpg or vice versa."""
    if not image_file_name:
        return None
    p = IMG_DIR / image_file_name
    if p.exists():
        return p
    # try other extensions
    stem = Path(image_file_name).stem
    for ext in (".jpeg", ".jpg", ".png", ".webp"):
        alt = IMG_DIR / f"{stem}{ext}"
        if alt.exists():
            return alt
    return None


# ---------------------------------------------------------------------------
# Supabase
# ---------------------------------------------------------------------------

sb: Client = create_client(SUPABASE_URL, SERVICE_KEY)


def get_or_create_user(email: str, display_name: str) -> str:
    """Return the auth.users.id for `email`, creating the user if needed."""
    # Page through existing users and look for a match.
    page = 1
    while True:
        resp = sb.auth.admin.list_users(page=page, per_page=200)
        # supabase-py returns either a list, or an object with .users — normalize.
        users = resp if isinstance(resp, list) else getattr(resp, "users", []) or []
        for u in users:
            if (u.email or "").lower() == email.lower():
                return u.id
        if not users or len(users) < 200:
            break
        page += 1

    # Not found — create.
    new_user = sb.auth.admin.create_user({
        "email": email,
        "email_confirm": True,
        "user_metadata": {"display_name": display_name},
    })
    user_obj = getattr(new_user, "user", None) or new_user
    return user_obj.id


def update_profile(user_id: str, fields: dict):
    payload = {k: v for k, v in fields.items() if v is not None}
    if not payload:
        return
    sb.table("profiles").update(payload).eq("user_id", user_id).execute()


def upload_image(local_path: Path, dest_path: str) -> None:
    mime, _ = mimetypes.guess_type(local_path.name)
    mime = mime or "application/octet-stream"
    with open(local_path, "rb") as f:
        data = f.read()
    sb.storage.from_(BUCKET).upload(
        path=dest_path,
        file=data,
        file_options={"content-type": mime, "upsert": "true"},
    )


def wipe_test_seeds():
    print("Wiping test seed listings…")
    sb.table("listings").delete().in_("listing_id", TEST_SEED_UUIDS).execute()


def wipe_legacy_imports():
    """If we previously ran this script, wipe its inserts so re-runs are clean.
    We mark imported rows by setting moderation_notes='legacy_import' so we
    can find them later. (Other rows are untouched.)
    """
    print("Wiping any prior legacy_import listings…")
    sb.table("listings").delete().eq("moderation_notes", "legacy_import").execute()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not XLSX_PATH.exists():
        sys.exit(f"ERROR: spreadsheet not found at {XLSX_PATH}")
    if not IMG_DIR.exists():
        sys.exit(f"ERROR: image directory not found at {IMG_DIR}")

    print(f"Reading {XLSX_PATH.name}…")
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Form Responses 1"]
    rows = [r for r in ws.iter_rows(values_only=True)][1:]
    rows = [r for r in rows if any(r)]
    print(f"  {len(rows)} non-empty rows")

    # ---- Step 1: wipe seed/prior imports ----
    wipe_test_seeds()
    wipe_legacy_imports()

    # ---- Step 2: collect per-seller data ----
    sellers: dict[str, dict] = {}
    for r in rows:
        email = normalize_email(r[COL["seller_email_pref"]] or r[COL["seller_email_alt"]])
        if not email:
            continue
        s = sellers.setdefault(email, {
            "display_name": None, "facebook_profile_url": None,
            "post_header": None, "post_footer": None, "location": None,
        })
        # Most-recent non-empty wins
        for key, idx in [
            ("display_name", COL["seller_name"]),
            ("facebook_profile_url", COL["seller_profile_url"]),
            ("post_header", COL["seller_post_header"]),
            ("post_footer", COL["seller_post_footer"]),
            ("location", COL["seller_location"]),
        ]:
            v = r[idx]
            if v:
                s[key] = v

    # ---- Step 3: ensure auth users + update profiles ----
    user_ids: dict[str, str] = {}
    for email, info in sellers.items():
        print(f"Upserting user {email}…")
        user_ids[email] = get_or_create_user(email, info["display_name"] or email)
        update_profile(user_ids[email], info)

    # Mark Guy as admin (he's the catalog runner).
    if "gjscuderi@gmail.com" in user_ids:
        sb.table("profiles").update({"is_admin": True}) \
            .eq("user_id", user_ids["gjscuderi@gmail.com"]).execute()

    # Mark trusted sellers from the second sheet.
    if "Trusted Sellers" in wb.sheetnames:
        ts = wb["Trusted Sellers"]
        for row in ts.iter_rows(min_row=2, values_only=True):
            email = normalize_email(row[0])
            if email and email in user_ids:
                sb.table("profiles").update({"is_trusted": True}) \
                    .eq("user_id", user_ids[email]).execute()

    # ---- Step 4: per-listing import ----
    inserted = 0
    skipped  = 0
    image_misses: list[str] = []
    for r in rows:
        email = normalize_email(r[COL["seller_email_pref"]] or r[COL["seller_email_alt"]])
        if not email or email not in user_ids:
            skipped += 1
            continue

        h, w, d = parse_dimensions(r[COL["size"]])
        new_uuid = str(uuid.uuid4())

        listing = {
            "listing_id":         new_uuid,
            "seller_id":          user_ids[email],
            "artist_name":        (r[COL["artist_name"]] or "Unknown").strip(),
            "artwork_title":      (r[COL["artwork_title"]] or "Untitled").strip(),
            "artwork_category":   r[COL["category"]] or "Other",
            "medium":             r[COL["medium"]],
            "height_in":          h,
            "width_in":           w,
            "depth_in":           d,
            "framed_size":        r[COL["framed_size"]],
            "description":        r[COL["description"]],
            "asking_price_usd":   float(r[COL["price"]] or 0),
            "previous_price_usd": float(r[COL["previous_price"]]) if r[COL["previous_price"]] else None,
            "price_updated_at":   iso_or_none(r[COL["price_updated_at"]]),
            "shipping_offered":   r[COL["shipping"]] or "No",
            "coa_included":       r[COL["coa"]] or "No",
            "status":             r[COL["status"]] or "available",
            "moderation_status":  r[COL["moderation_status"]] or "approved",
            "moderation_notes":   "legacy_import",
            "seller_mood":        r[COL["seller_mood"]],
            "last_renewed_at":    iso_or_none(r[COL["last_renewed_at"]]) or iso_or_none(r[COL["timestamp"]]),
            "created_at":         iso_or_none(r[COL["timestamp"]]),
        }

        # Upload image (if found)
        image_file = r[COL["image_file"]]
        local = find_image_file(image_file) if image_file else None
        storage_path = None
        if local:
            storage_path = f"{new_uuid}/0-{local.name}"
            try:
                upload_image(local, storage_path)
            except Exception as e:
                print(f"  ! image upload failed for {image_file}: {e}")
                storage_path = None
        elif image_file:
            image_misses.append(image_file)

        sb.table("listings").insert(listing).execute()
        if storage_path:
            sb.table("listing_images").insert({
                "listing_id":   new_uuid,
                "storage_path": storage_path,
                "position":     0,
            }).execute()

        inserted += 1
        legacy_id = r[COL["listing_id"]]
        print(f"  + {legacy_id} → {new_uuid[:8]}…  {listing['artist_name']} — {listing['artwork_title']}")

    # ---- Done ----
    print()
    print(f"Inserted: {inserted}")
    print(f"Skipped:  {skipped}")
    if image_misses:
        print(f"Image files referenced but not found in catalog_images/:")
        for m in image_misses:
            print(f"  - {m}")


if __name__ == "__main__":
    main()
